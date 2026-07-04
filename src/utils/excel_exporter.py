import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Styles ─────────────────────────────────────────────────────────────────────
_HEADER_FILL   = PatternFill("solid", fgColor="1A5276")
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FILL  = PatternFill("solid", fgColor="2E86C1")
_SECTION_FONT  = Font(bold=True, color="FFFFFF", size=12)
_OUTLIER_FILL  = PatternFill("solid", fgColor="FADBD8")
_ALT_FILL      = PatternFill("solid", fgColor="EBF5FB")
_CHANGED_FILL  = PatternFill("solid", fgColor="D5F5E3")
_WARN_FILL     = PatternFill("solid", fgColor="FEF9E7")
_LABEL_FONT    = Font(bold=True)
_CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT          = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _auto_width(ws, max_width: int = 45) -> None:
    for col in ws.columns:
        width = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, max_width)


def _section_header(ws, row: int, text: str, span: int = 6) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = _SECTION_FILL
    cell.font = _SECTION_FONT
    cell.alignment = _LEFT
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def _write_header_row(ws, row: int, headers: list[str]) -> int:
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
    return row + 1


def _safe_val(val):
    if hasattr(val, "item"):
        return val.item()
    if pd.isna(val) if not isinstance(val, (list, dict)) else False:
        return None
    return val


def _write_df(ws, df: pd.DataFrame, outlier_indices: set | None = None, start_row: int = 1) -> int:
    row = _write_header_row(ws, start_row, df.columns.tolist())
    outlier_set = outlier_indices or set()
    for orig_idx, data_row in df.iterrows():
        is_outlier = orig_idx in outlier_set
        for col_idx, val in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col_idx, value=_safe_val(val))
            cell.alignment = _LEFT
            if is_outlier:
                cell.fill = _OUTLIER_FILL
            elif (row - start_row) % 2 == 1:
                cell.fill = _ALT_FILL
        row += 1
    return row


def generate_excel_report(result: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    cleaned_df: pd.DataFrame = result.get("cleaned_df")
    eda          = result.get("eda_report", {})
    outlier_rep  = result.get("outlier_report", {})
    outlier_cols = outlier_rep.get("columns", {})
    dtypes_rep   = result.get("dtypes_report", {})
    column_map   = result.get("column_map", {})
    narrative    = result.get("final_report", {}).get("narrative", "")

    all_outlier_idx: set = set()
    for col_data in outlier_cols.values():
        all_outlier_idx.update(col_data.get("iqr_outlier_indices", []))

    # ── Sheet 1: Cleaned Data ──────────────────────────────────────────────────
    if cleaned_df is not None:
        ws1 = wb.create_sheet("Cleaned Data")
        _write_df(ws1, cleaned_df, outlier_indices=all_outlier_idx)
        ws1.freeze_panes = "A2"
        ws1.auto_filter.ref = ws1.dimensions
        _auto_width(ws1)

        legend_row = len(cleaned_df) + 3
        ws1.cell(row=legend_row, column=1, value="Legend:").font = _LABEL_FONT
        ws1.cell(row=legend_row, column=2, value="  Outlier row (IQR)  ").fill  = _OUTLIER_FILL
        ws1.cell(row=legend_row, column=3, value="  Alternating row  ").fill    = _ALT_FILL

    # ── Sheet 2: EDA Summary ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("EDA Summary")
    r = 1

    if narrative:
        r = _section_header(ws2, r, "AI Narrative", span=6)
        cell = ws2.cell(row=r, column=1, value=narrative)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[r].height = 130
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 2

    shape = eda.get("shape", {})
    if shape:
        r = _section_header(ws2, r, "Dataset Shape")
        for label, val in [("Rows", shape.get("rows")), ("Columns", shape.get("columns"))]:
            ws2.cell(row=r, column=1, value=label).font = _LABEL_FONT
            ws2.cell(row=r, column=2, value=val)
            r += 1
        r += 1

    null_pct = eda.get("null_pct", {})
    if null_pct:
        r = _section_header(ws2, r, "Null % per Column", span=2)
        r = _write_header_row(ws2, r, ["Column", "Null %"])
        for col, pct in sorted(null_pct.items(), key=lambda x: -x[1]):
            ws2.cell(row=r, column=1, value=col)
            cell = ws2.cell(row=r, column=2, value=round(pct, 2))
            if pct > 20:
                cell.fill = _OUTLIER_FILL
            elif pct > 0:
                cell.fill = _WARN_FILL
            r += 1
        r += 1

    summary = eda.get("summary_stats", {})
    if summary:
        r = _section_header(ws2, r, "Summary Statistics", span=10)
        stats_df = pd.DataFrame(summary).T.reset_index().rename(columns={"index": "Column"})
        r = _write_header_row(ws2, r, stats_df.columns.tolist())
        for _, row_data in stats_df.iterrows():
            for col_idx, val in enumerate(row_data, 1):
                v = round(float(val), 4) if col_idx > 1 else val
                ws2.cell(row=r, column=col_idx, value=v).alignment = _LEFT
            r += 1

    _auto_width(ws2)
    ws2.column_dimensions["A"].width = 28

    # ── Sheet 3: Outlier Report ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Outlier Report")
    r = 1
    r = _section_header(ws3, r, "Outlier Summary by Column", span=5)
    r = _write_header_row(ws3, r, ["Column", "IQR Outliers", "Z-Score Outliers", "Lower Bound", "Upper Bound"])
    for col_name, data in outlier_cols.items():
        vals = [col_name, data["iqr_outlier_count"], data["z_score_outlier_count"],
                data["bounds"]["lower"], data["bounds"]["upper"]]
        for col_idx, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=col_idx, value=v)
            if data["iqr_outlier_count"] > 0:
                cell.fill = _OUTLIER_FILL
        r += 1
    r += 1

    if cleaned_df is not None and all_outlier_idx:
        r = _section_header(ws3, r, f"Flagged Rows ({len(all_outlier_idx)} total)", span=len(cleaned_df.columns))
        flagged = cleaned_df.loc[sorted(all_outlier_idx)].reset_index(drop=True)
        _write_df(ws3, flagged, start_row=r)

    _auto_width(ws3)

    # ── Sheet 4: Data Changes ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("Data Changes")
    r = 1
    r = _section_header(ws4, r, "Header Mapping", span=2)
    r = _write_header_row(ws4, r, ["Original Header", "Cleaned Header"])
    for orig, clean in column_map.items():
        ws4.cell(row=r, column=1, value=orig).fill = (_WARN_FILL if orig != clean else PatternFill())
        ws4.cell(row=r, column=2, value=clean).fill = (_CHANGED_FILL if orig != clean else PatternFill())
        r += 1
    r += 2

    r = _section_header(ws4, r, "Type Inference", span=4)
    r = _write_header_row(ws4, r, ["Column", "Original Type", "Inferred Type", "Action"])
    for col_name, info in dtypes_rep.items():
        changed = info["action"] != "unchanged"
        for col_idx, v in enumerate([col_name, info["original"], info["inferred"], info["action"]], 1):
            cell = ws4.cell(row=r, column=col_idx, value=v)
            if changed:
                cell.fill = _CHANGED_FILL
        r += 1

    _auto_width(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
